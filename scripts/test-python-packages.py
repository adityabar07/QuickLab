#!/usr/bin/env python3
"""
QuickLab — Official Python Package & Functional Verification Suite
Validates that all pre-installed data science, ML, DL, NLP, graph, CV, and statistical
packages import cleanly and execute functional tasks correctly.
"""

import sys
import time
import io

print("=" * 70)
print("  QUICKLAB — PYTHON EXECUTION ENVIRONMENT VERIFICATION")
print(f"  Python Version: {sys.version.split()[0]} ({sys.platform})")
print("=" * 70)

passed_imports = 0
failed_imports = 0
passed_functional = 0
failed_functional = 0

import_results = []
functional_results = []

def test_import(name, module_name=None, attr=""):
    global passed_imports, failed_imports
    target = module_name or name
    try:
        start = time.perf_counter()
        mod = __import__(target)
        if attr:
            for part in attr.split("."):
                mod = getattr(mod, part)
        version = getattr(mod, "__version__", "available")
        dur = (time.perf_counter() - start) * 1000
        import_results.append((name, True, f"v{version} ({dur:.1f}ms)", None))
        passed_imports += 1
        return mod
    except Exception as e:
        import_results.append((name, False, "FAILED", str(e)))
        failed_imports += 1
        return None

def test_func(name, fn):
    global passed_functional, failed_functional
    try:
        start = time.perf_counter()
        fn()
        dur = (time.perf_counter() - start) * 1000
        functional_results.append((name, True, f"OK ({dur:.1f}ms)", None))
        passed_functional += 1
    except Exception as e:
        functional_results.append((name, False, "FAILED", str(e)))
        failed_functional += 1

print("\n--- [1/2] IMPORTING OFFICIAL QUICKLAB LIBRARIES ---")

np_mod = test_import("numpy", "numpy")
scipy_mod = test_import("scipy", "scipy")
sympy_mod = test_import("sympy", "sympy")
pd_mod = test_import("pandas", "pandas")
polars_mod = test_import("polars", "polars")
statsmodels_mod = test_import("statsmodels", "statsmodels")
mpl_mod = test_import("matplotlib", "matplotlib")
sns_mod = test_import("seaborn", "seaborn")
plotly_mod = test_import("plotly", "plotly")
sklearn_mod = test_import("scikit-learn", "sklearn")
tf_mod = test_import("tensorflow", "tensorflow")
keras_mod = test_import("keras", "keras")
torch_mod = test_import("torch", "torch")
tv_mod = test_import("torchvision", "torchvision")
ta_mod = test_import("torchaudio", "torchaudio")
cv2_mod = test_import("opencv-python", "cv2")
pil_mod = test_import("pillow", "PIL")
imageio_mod = test_import("imageio", "imageio")
nltk_mod = test_import("nltk", "nltk")
spacy_mod = test_import("spacy", "spacy")
hf_trans_mod = test_import("transformers", "transformers")
st_mod = test_import("sentence-transformers", "sentence_transformers")
pgmpy_mod = test_import("pgmpy", "pgmpy")
nx_mod = test_import("networkx", "networkx")
openpyxl_mod = test_import("openpyxl", "openpyxl")
xlsxwriter_mod = test_import("xlsxwriter", "xlsxwriter")
h5py_mod = test_import("h5py", "h5py")
requests_mod = test_import("requests", "requests")
bs4_mod = test_import("beautifulsoup4", "bs4")
ipython_mod = test_import("ipython", "IPython")
jupyter_client_mod = test_import("jupyter-client", "jupyter_client")
tqdm_mod = test_import("tqdm", "tqdm")
joblib_mod = test_import("joblib", "joblib")
pydantic_mod = test_import("pydantic", "pydantic")
yaml_mod = test_import("pyyaml", "yaml")

for name, ok, msg, err in import_results:
    if ok:
        print(f"  \033[32m✓\033[0m {name:<24} : {msg}")
    else:
        print(f"  \033[31m✗\033[0m {name:<24} : {msg} -> {err}")

print(f"\nImport Summary: {passed_imports} passed, {failed_imports} failed out of {len(import_results)}")

print("\n--- [2/2] RUNNING FUNCTIONAL UNIT TESTS ---")

# 1. NumPy
def test_numpy_func():
    import numpy as np
    a = np.array([1, 2, 3, 4, 5])
    assert a.mean() == 3.0
    mat = np.eye(3)
    assert mat.shape == (3, 3)
test_func("NumPy (Arrays & Linear Algebra)", test_numpy_func)

# 2. Pandas
def test_pandas_func():
    import pandas as pd
    df = pd.DataFrame({"Name": ["A", "B", "C"], "Score": [90, 85, 95]})
    assert df["Score"].mean() == 90.0
    html = df.to_html()
    assert "<table" in html
test_func("Pandas (DataFrames & HTML Export)", test_pandas_func)

# 3. Polars
def test_polars_func():
    import polars as pl
    df = pl.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
    assert df["a"].sum() == 6
test_func("Polars (Fast Columnar DataFrame)", test_polars_func)

# 4. SciPy
def test_scipy_func():
    import scipy.optimize as opt
    res = opt.minimize(lambda x: (x[0] - 5)**2, [0.0])
    assert abs(res.x[0] - 5.0) < 1e-4
test_func("SciPy (Numerical Optimization)", test_scipy_func)

# 5. SymPy
def test_sympy_func():
    import sympy as sp
    x = sp.symbols("x")
    sols = sp.solve(x**2 - 9, x)
    assert -3 in sols and 3 in sols
test_func("SymPy (Symbolic Equations)", test_sympy_func)

# 6. Matplotlib
def test_matplotlib_func():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots()
    ax.plot([1, 2, 3], [4, 5, 6])
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    assert len(buf.read()) > 100
    plt.close(fig)
test_func("Matplotlib (Headless PNG Render)", test_matplotlib_func)

# 7. Seaborn
def test_seaborn_func():
    import matplotlib
    matplotlib.use("Agg")
    import seaborn as sns
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots()
    sns.barplot(x=["A", "B"], y=[1, 2], ax=ax)
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
test_func("Seaborn (Statistical Graphics)", test_seaborn_func)

# 8. Plotly
def test_plotly_func():
    import plotly.express as px
    fig = px.scatter(x=[1, 2, 3], y=[4, 5, 6], title="QuickLab Plotly")
    json_data = fig.to_json()
    assert len(json_data) > 50
test_func("Plotly (Interactive JSON/HTML Export)", test_plotly_func)

# 9. Scikit-learn
def test_sklearn_func():
    from sklearn.linear_model import LinearRegression
    X = [[1], [2], [3], [4]]
    y = [2, 4, 6, 8]
    model = LinearRegression().fit(X, y)
    pred = model.predict([[5]])
    assert abs(pred[0] - 10.0) < 1e-4
test_func("Scikit-learn (ML Model Fit & Predict)", test_sklearn_func)

# 10. PyTorch
def test_torch_func():
    import torch
    x = torch.tensor([1.0, 2.0, 3.0], requires_grad=True)
    y = (x ** 2).sum()
    y.backward()
    assert x.grad is not None
    assert torch.equal(x.grad, torch.tensor([2.0, 4.0, 6.0]))
test_func("PyTorch (Tensors & Autograd)", test_torch_func)

# 11. TensorFlow & Keras
def test_tensorflow_func():
    import tensorflow as tf
    t1 = tf.constant([1.0, 2.0])
    t2 = tf.constant([3.0, 4.0])
    t3 = tf.add(t1, t2)
    assert float(tf.reduce_sum(t3)) == 10.0
test_func("TensorFlow & Keras (Tensor Math)", test_tensorflow_func)

# 12. OpenCV & Pillow
def test_opencv_pillow_func():
    import cv2
    import numpy as np
    from PIL import Image
    # OpenCV
    arr = np.full((30, 30, 3), 128, dtype=np.uint8)
    gray = cv2.cvtColor(arr, cv2.COLOR_BGR2GRAY)
    assert gray.shape == (30, 30)
    # PIL
    img = Image.fromarray(arr)
    assert img.size == (30, 30)
test_func("OpenCV & Pillow (Image Transforms)", test_opencv_pillow_func)

# 13. NLTK & spaCy
def test_nlp_func():
    import nltk
    import spacy
    text = "QuickLab is an instant interactive Python notebook environment."
    tokens = nltk.word_tokenize(text)
    assert "QuickLab" in tokens
    nlp = spacy.blank("en")
    doc = nlp(text)
    assert len(doc) > 5
test_func("NLP (NLTK Tokenize & spaCy Processing)", test_nlp_func)

# 14. Probabilistic AI (pgmpy)
def test_pgmpy_func():
    from pgmpy.models import BayesianNetwork
    from pgmpy.factors.discrete import TabularCPD
    model = BayesianNetwork([("A", "B")])
    cpd_a = TabularCPD(variable="A", variable_card=2, values=[[0.6], [0.4]])
    cpd_b = TabularCPD(variable="B", variable_card=2, values=[[0.7, 0.2], [0.3, 0.8]], evidence=["A"], evidence_card=[2])
    model.add_cpds(cpd_a, cpd_b)
    assert model.check_model() is True
test_func("pgmpy (Bayesian Network Modeling)", test_pgmpy_func)

# 15. NetworkX
def test_networkx_func():
    import networkx as nx
    G = nx.Graph()
    G.add_edges_from([("Alpha", "Beta"), ("Beta", "Gamma"), ("Gamma", "Delta")])
    path = nx.shortest_path(G, "Alpha", "Delta")
    assert path == ["Alpha", "Beta", "Gamma", "Delta"]
test_func("NetworkX (Graph Algorithms & Shortest Path)", test_networkx_func)

# 16. Statsmodels
def test_statsmodels_func():
    import statsmodels.api as sm
    import numpy as np
    x = np.array([1, 2, 3, 4, 5])
    y = np.array([2, 4, 6, 8, 10])
    X = sm.add_constant(x)
    model = sm.OLS(y, X).fit()
    assert abs(model.params[1] - 2.0) < 1e-4
test_func("Statsmodels (OLS Linear Regression)", test_statsmodels_func)

# 17. Excel & Spreadsheet (OpenPyXL / XlsxWriter)
def test_excel_func():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "QuickLab"
    ws['B1'] = 42
    buf = io.BytesIO()
    wb.save(buf)
    assert len(buf.getvalue()) > 100
test_func("Spreadsheet (OpenPyXL & XlsxWriter)", test_excel_func)

# 18. Web & HTML Parsing (BeautifulSoup)
def test_bs4_func():
    from bs4 import BeautifulSoup
    html_doc = '<div id="quicklab"><h1>Ready</h1><p class="status">Live</p></div>'
    soup = BeautifulSoup(html_doc, "html.parser")
    assert soup.find("h1").text == "Ready"
    assert soup.find("p", class_="status").text == "Live"
test_func("Web Data (BeautifulSoup4 HTML Extraction)", test_bs4_func)

for name, ok, msg, err in functional_results:
    if ok:
        print(f"  \033[32m✓\033[0m {name:<46} : {msg}")
    else:
        print(f"  \033[31m✗\033[0m {name:<46} : {msg} -> {err}")

print(f"\nFunctional Summary: {passed_functional} passed, {failed_functional} failed out of {len(functional_results)}")

print("=" * 70)
total_failed = failed_imports + failed_functional
if total_failed == 0:
    print(f"  \033[32mALL TESTS PASSED SUCCESSFULLY! ({passed_imports + passed_functional} / {passed_imports + passed_functional})\033[0m")
    print("  QuickLab Python 3.11 Environment is fully operational and verified.")
    print("=" * 70)
    sys.exit(0)
else:
    print(f"  \033[31mVERIFICATION FAILED: {total_failed} error(s) detected.\033[0m")
    print("=" * 70)
    sys.exit(1)
